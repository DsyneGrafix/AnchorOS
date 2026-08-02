"""AOS-180 — AnchorOS Data Exchange Service.

Platform-wide, application-neutral import/export service for tabular records.
Supports CSV, JSON, and Excel (XLSX). Every exchange produces a structured
receipt with provenance and integrity metadata.
"""
from __future__ import annotations

import csv
import hashlib
import json
from copy import copy
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence
from uuid import uuid4

from core.module import Module

try:
    from services.event import AnchorEvent
except ImportError:  # Standalone verification fallback
    AnchorEvent = None  # type: ignore[assignment,misc]

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - exercised only without dependency
    Workbook = None  # type: ignore[assignment,misc]
    load_workbook = None  # type: ignore[assignment,misc]


Record = dict[str, Any]


class DataExchangeError(RuntimeError):
    """Base error for failed data-exchange operations."""


class UnsupportedFormatError(DataExchangeError):
    """Raised when an exchange format is unsupported."""


class DataValidationError(DataExchangeError):
    """Raised when imported or exported records are invalid."""


@dataclass(frozen=True, slots=True)
class ExchangeReceipt:
    """Immutable provenance record for one import or export operation."""

    exchange_id: str
    direction: str
    format: str
    source: str
    destination: str
    record_count: int
    sha256: str
    generated_at: str
    service: str = "AOS-180 Data Exchange Service"
    service_version: str = "1.0.0"

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


class DataExchangeService(Module):
    """Application-neutral AnchorOS data import/export capability."""

    SUPPORTED_FORMATS = {"csv", "json", "xlsx"}

    def __init__(self, context: Any | None = None) -> None:
        super().__init__("Data Exchange Service", "1.0.0")
        self._context = context
        self._receipts: list[ExchangeReceipt] = []

    def supported_formats(self) -> tuple[str, ...]:
        return tuple(sorted(self.SUPPORTED_FORMATS))

    def get_receipts(self) -> list[dict[str, Any]]:
        return [receipt.to_dict() for receipt in self._receipts]

    def import_records(
        self,
        source: str | Path,
        *,
        format_name: str | None = None,
        sheet_name: str | None = None,
        required_fields: Sequence[str] | None = None,
    ) -> tuple[list[Record], ExchangeReceipt]:
        """Import records and return normalized dictionaries plus a receipt."""
        path = self._safe_path(source, must_exist=True)
        fmt = self._resolve_format(path, format_name)

        if fmt == "csv":
            records = self._import_csv(path)
        elif fmt == "json":
            records = self._import_json(path)
        elif fmt == "xlsx":
            records = self._import_xlsx(path, sheet_name=sheet_name)
        else:  # defensive; _resolve_format already checks
            raise UnsupportedFormatError(fmt)

        normalized = self._normalize_records(records)
        self._validate_required_fields(normalized, required_fields)
        receipt = self._record_exchange(
            direction="IMPORT",
            format_name=fmt,
            source=str(path),
            destination="AnchorOS runtime",
            record_count=len(normalized),
            artifact=path,
        )
        return normalized, receipt

    def export_records(
        self,
        records: Iterable[Mapping[str, Any]],
        destination: str | Path,
        *,
        format_name: str | None = None,
        sheet_name: str = "Data",
        required_fields: Sequence[str] | None = None,
        write_receipt: bool = True,
    ) -> ExchangeReceipt:
        """Export normalized records and return an immutable receipt."""
        path = self._safe_path(destination, must_exist=False)
        fmt = self._resolve_format(path, format_name)
        normalized = self._normalize_records(records)
        self._validate_required_fields(normalized, required_fields)
        path.parent.mkdir(parents=True, exist_ok=True)

        if fmt == "csv":
            self._export_csv(normalized, path)
        elif fmt == "json":
            self._export_json(normalized, path)
        elif fmt == "xlsx":
            self._export_xlsx(normalized, path, sheet_name=sheet_name)
        else:
            raise UnsupportedFormatError(fmt)

        receipt = self._record_exchange(
            direction="EXPORT",
            format_name=fmt,
            source="AnchorOS runtime",
            destination=str(path),
            record_count=len(normalized),
            artifact=path,
        )
        if write_receipt:
            receipt_path = path.with_suffix(path.suffix + ".receipt.json")
            receipt_path.write_text(
                json.dumps(receipt.to_dict(), indent=2, sort_keys=True),
                encoding="utf-8",
            )
        return receipt

    def health(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "version": self.version,
            "status": self.status,
            "supported_formats": list(self.supported_formats()),
            "receipt_count": len(self._receipts),
            "xlsx_available": Workbook is not None,
        }

    @staticmethod
    def _safe_path(value: str | Path, *, must_exist: bool) -> Path:
        path = Path(value).expanduser().resolve()
        if must_exist and not path.is_file():
            raise DataExchangeError(f"Source file does not exist: {path}")
        return path

    def _resolve_format(self, path: Path, format_name: str | None) -> str:
        fmt = (format_name or path.suffix.lstrip(".")).lower().strip()
        if fmt == "xls":
            raise UnsupportedFormatError("Legacy .xls is not supported; use .xlsx")
        if fmt not in self.SUPPORTED_FORMATS:
            raise UnsupportedFormatError(
                f"Unsupported format '{fmt}'. Supported: {', '.join(self.supported_formats())}"
            )
        return fmt

    @staticmethod
    def _normalize_records(records: Iterable[Mapping[str, Any]]) -> list[Record]:
        normalized: list[Record] = []
        for index, record in enumerate(records, start=1):
            if not isinstance(record, Mapping):
                raise DataValidationError(f"Record {index} is not a mapping")
            clean = {str(key).strip(): value for key, value in record.items() if str(key).strip()}
            if not clean:
                raise DataValidationError(f"Record {index} is empty")
            normalized.append(clean)
        return normalized

    @staticmethod
    def _validate_required_fields(
        records: Sequence[Record], required_fields: Sequence[str] | None
    ) -> None:
        if not required_fields:
            return
        required = {field.strip() for field in required_fields if field.strip()}
        for index, record in enumerate(records, start=1):
            missing = sorted(field for field in required if field not in record or record[field] in (None, ""))
            if missing:
                raise DataValidationError(
                    f"Record {index} missing required fields: {', '.join(missing)}"
                )

    @staticmethod
    def _fieldnames(records: Sequence[Record]) -> list[str]:
        fields: list[str] = []
        seen: set[str] = set()
        for record in records:
            for key in record:
                if key not in seen:
                    seen.add(key)
                    fields.append(key)
        return fields

    @staticmethod
    def _import_csv(path: Path) -> list[Record]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise DataValidationError("CSV file has no header row")
            return [dict(row) for row in reader]

    @staticmethod
    def _export_csv(records: Sequence[Record], path: Path) -> None:
        fields = DataExchangeService._fieldnames(records)
        if not fields:
            raise DataValidationError("Cannot export an empty record set")
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(records)

    @staticmethod
    def _import_json(path: Path) -> list[Record]:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, dict) and "records" in payload:
            payload = payload["records"]
        if not isinstance(payload, list):
            raise DataValidationError("JSON input must be an array or an object containing 'records'")
        return payload

    @staticmethod
    def _export_json(records: Sequence[Record], path: Path) -> None:
        path.write_text(
            json.dumps({"records": list(records)}, indent=2, default=str),
            encoding="utf-8",
        )

    @staticmethod
    def _require_xlsx() -> None:
        if Workbook is None or load_workbook is None:
            raise DataExchangeError(
                "Excel support requires openpyxl. Install with: pip install openpyxl>=3.1"
            )

    @classmethod
    def _import_xlsx(cls, path: Path, *, sheet_name: str | None) -> list[Record]:
        cls._require_xlsx()
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            rows = worksheet.iter_rows(values_only=True)
            try:
                raw_headers = next(rows)
            except StopIteration as exc:
                raise DataValidationError("Excel worksheet is empty") from exc
            headers = [str(value).strip() if value is not None else "" for value in raw_headers]
            if not any(headers):
                raise DataValidationError("Excel worksheet has no header row")
            records: list[Record] = []
            for values in rows:
                if all(value is None for value in values):
                    continue
                records.append({headers[i]: value for i, value in enumerate(values) if i < len(headers) and headers[i]})
            return records
        finally:
            workbook.close()

    @classmethod
    def _export_xlsx(cls, records: Sequence[Record], path: Path, *, sheet_name: str) -> None:
        cls._require_xlsx()
        fields = cls._fieldnames(records)
        if not fields:
            raise DataValidationError("Cannot export an empty record set")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = (sheet_name or "Data")[:31]
        worksheet.append(fields)
        for record in records:
            worksheet.append([record.get(field) for field in fields])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 50)
        workbook.save(path)

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def _record_exchange(
        self,
        *,
        direction: str,
        format_name: str,
        source: str,
        destination: str,
        record_count: int,
        artifact: Path,
    ) -> ExchangeReceipt:
        receipt = ExchangeReceipt(
            exchange_id=f"AOS-180-{direction[:3]}-{datetime.now(timezone.utc):%Y%m%d}-{uuid4().hex[:8].upper()}",
            direction=direction,
            format=format_name.upper(),
            source=source,
            destination=destination,
            record_count=record_count,
            sha256=self._sha256(artifact),
            generated_at=datetime.now(timezone.utc).isoformat(),
        )
        self._receipts.append(receipt)
        self._publish_event(receipt)
        return receipt

    def _publish_event(self, receipt: ExchangeReceipt) -> None:
        if AnchorEvent is None or self._context is None:
            return
        try:
            event_bus = self._context.require("Event Bus")
        except (AttributeError, KeyError, RuntimeError):
            return
        event_bus.publish(
            AnchorEvent(
                source="AOS-180",
                event_type=f"data_exchange.{receipt.direction.lower()}.completed",
                message=f"{receipt.direction.title()} completed for {receipt.record_count} records.",
                payload=receipt.to_dict(),
            )
        )


def create_module(context: Any) -> DataExchangeService:
    """AnchorOS discovery factory."""
    return DataExchangeService(context=context)
